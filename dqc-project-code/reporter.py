import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side
import os


'''
Method to print a summary of the quality check to the console
'''
def print_result(result):
    print(f"[{result["qc_name"]}]:\t{result["status"]}")
    
    # only need to print details if quality check failed
    if result["status"] == "FAIL":
        print("Details:")
        for table, details in result["details"].items():
            print(table.upper() + ": ")

            if isinstance(details, list):
                for row in details:
                    print(row)
            else:
                print(details)

    print("-" * 60)


'''
Method to save all quality check results ot a timestamped xlsx file
'''
def save_results_xlsx(results, output_dir="results"):
    os.makedirs(output_dir, exist_ok=True) # create folder
    output_path = os.path.join(output_dir, "dqf_results_testing.xlsx")

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        for table, results in results.items():
            rows = []
            for r in results:
                row = {
                    "Quality Check Name": r["qc_name"],
                    "Status": r["status"]
                }

                # fill up details cells
                details_cell = []
                for key, val in r["details"].items():
                    if isinstance(val, list):
                        details_list = ("\n").join(str(item) for item in val)
                        cell_info = f"{key}: \n{details_list}"
                    else:
                        cell_info = f"{key}: \n{str(val)}"

                    details_cell.append(cell_info)

                row["Details"] = ("\n\n").join(str(item) for item in details_cell)

                rows.append(row)

            df = pd.DataFrame(rows) # write all the rows
            df.to_excel(excel_writer=writer, sheet_name=table.upper(), index=False) # create sheet

    # style workbook
    wb = load_workbook(output_path)
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        ws.freeze_panes = "A2" # freeze top row

        # set column widths
        ws.column_dimensions["A"].width = 40 # quality check name
        ws.column_dimensions["B"].width = 12 # status
        ws.column_dimensions["C"].width = 100 # results

        # format columns A and B to justify text at the top of the cell
        for row in range(2, ws.max_row + 1):
            for col in ["A", "B"]:
                cell = ws[f"{col}{row}"]
                cell.alignment = cell.alignment.copy(vertical="top")
        
        # apply conditional formatting for Status column
        for row in range(2, ws.max_row + 1):
            cell = ws[f"B{row}"]
            value = str(cell.value).strip().upper()

            if value == "PASS":
                cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid") # green for PASS
            if value == "FAIL":
                cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid") # green for PASS

        # add thin borders
        thin_border = Border(
            left=Side(style="hair"),
            right=Side(style="hair"),
            top=Side(style="thin"),
            bottom=Side(style="thin")
        )

        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                cell.border = thin_border


    wb.save(output_path)

    print(f"Results saved to: {output_path}")